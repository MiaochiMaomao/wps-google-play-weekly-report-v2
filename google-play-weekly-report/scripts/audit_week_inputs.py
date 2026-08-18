from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


def normalize(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def parse_date(value: object, epoch) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, epoch).date()
        except Exception:
            return None
    text = str(value).strip().replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        pass
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def contains_placeholder(value: object) -> bool:
    if isinstance(value, str):
        return not value.strip() or bool(re.search(r"<[^>]+>", value))
    if isinstance(value, dict):
        return any(contains_placeholder(item) for item in value.values())
    if isinstance(value, list):
        return any(contains_placeholder(item) for item in value)
    return value is None


def require_config(config: dict) -> list[dict]:
    issues: list[dict] = []

    def need(path: str):
        value: object = config
        try:
            for part in path.split("."):
                value = value[part]  # type: ignore[index]
        except Exception:
            issues.append({"code": "config_value_missing", "path": path})
            return None
        if contains_placeholder(value):
            issues.append({"code": "config_value_incomplete", "path": path})
        return value

    for path in (
        "period.start",
        "period.end",
        "period.require_every_date",
        "workbook.all_reviews.sheet",
        "workbook.all_reviews.columns.version",
        "workbook.all_reviews.columns.date",
        "workbook.all_reviews.columns.rating",
        "workbook.all_reviews.columns.review",
        "workbook.classified_reviews.sheet",
        "workbook.classified_reviews.columns.date",
        "workbook.classified_reviews.columns.rating",
        "workbook.classified_reviews.columns.category",
        "workbook.classified_reviews.columns.original",
        "workbook.classified_reviews.columns.translation",
        "classification.categories",
        "classification.subcategory_columns",
        "classification.representative_category",
        "metrics.low_ratings",
        "metrics.high_ratings",
        "metrics.version_scope",
        "metrics.version_min_comments",
        "metrics.version_top_n",
        "metrics.version_sort",
        "metrics.issue_top_n",
        "metrics.representative_comment_min_chars",
        "metrics.representative_comment_max_chars",
        "format.score_decimals",
        "format.percentage_decimals",
        "format.unavailable_marker",
        "report.title",
        "report.output_filename",
    ):
        need(path)

    try:
        start = date.fromisoformat(config["period"]["start"])
        end = date.fromisoformat(config["period"]["end"])
        if end < start:
            issues.append({"code": "config_period_reversed"})
    except Exception:
        issues.append({"code": "config_period_invalid"})

    categories = config.get("classification", {}).get("categories", [])
    mappings = config.get("classification", {}).get("subcategory_columns", {})
    if not isinstance(categories, list) or not categories or len(categories) != len(set(categories)):
        issues.append({"code": "config_categories_invalid"})
    elif set(mappings) != set(categories):
        issues.append({"code": "config_subcategory_mapping_mismatch"})
    if config.get("classification", {}).get("representative_category") not in categories:
        issues.append({"code": "config_representative_category_invalid"})

    for field in ("low_ratings", "high_ratings"):
        values = config.get("metrics", {}).get(field, [])
        if not isinstance(values, list) or not values or any(type(v) is not int or not 1 <= v <= 5 for v in values):
            issues.append({"code": "config_rating_group_invalid", "field": field})
    if set(config.get("metrics", {}).get("low_ratings", [])) & set(config.get("metrics", {}).get("high_ratings", [])):
        issues.append({"code": "config_rating_groups_overlap"})
    if config.get("metrics", {}).get("version_scope") not in {"comments", "ratings"}:
        issues.append({"code": "config_version_scope_invalid"})
    for field in ("version_min_comments", "version_top_n", "issue_top_n"):
        value = config.get("metrics", {}).get(field)
        if type(value) is not int or value < 1:
            issues.append({"code": "config_positive_integer_required", "field": field})
    low = config.get("metrics", {}).get("representative_comment_min_chars")
    high = config.get("metrics", {}).get("representative_comment_max_chars")
    if type(low) is not int or type(high) is not int or low < 0 or high < low:
        issues.append({"code": "config_comment_length_invalid"})
    sort_fields = config.get("metrics", {}).get("version_sort", [])
    allowed = {"total", "low", "high", "low_rate", "high_rate", "version"}
    if not isinstance(sort_fields, list) or not sort_fields or any(str(v).lstrip("+-") not in allowed for v in sort_fields):
        issues.append({"code": "config_version_sort_invalid"})
    return issues


def header_map(sheet, requested: dict[str, str]) -> tuple[dict[str, int], list[str]]:
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    by_name = {normalize(value): index for index, value in enumerate(headers)}
    mapping: dict[str, int] = {}
    missing: list[str] = []
    for field, header in requested.items():
        index = by_name.get(normalize(header))
        if index is None:
            missing.append(header)
        else:
            mapping[field] = index
    return mapping, missing


def exact_row_key(values) -> tuple[str, ...]:
    return tuple("" if value is None else str(value).strip() for value in values)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--previous-report", required=True)
    parser.add_argument("--config", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    workbook_path = Path(args.workbook).resolve()
    previous_path = Path(args.previous_report).resolve()
    config_path = Path(args.config).resolve()
    output_path = Path(args.output).resolve()
    config = json.loads(config_path.read_text(encoding="utf-8"))
    issues = require_config(config)

    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    if not previous_path.exists():
        raise SystemExit(f"Previous report not found: {previous_path}")
    try:
        Document(previous_path)
    except Exception as exc:
        issues.append({"code": "previous_report_unreadable", "detail": str(exc)})

    wb = load_workbook(workbook_path, data_only=True, read_only=False)
    all_spec = config.get("workbook", {}).get("all_reviews", {})
    classified_spec = config.get("workbook", {}).get("classified_reviews", {})
    all_sheet = wb[all_spec["sheet"]] if all_spec.get("sheet") in wb.sheetnames else None
    classified_sheet = wb[classified_spec["sheet"]] if classified_spec.get("sheet") in wb.sheetnames else None
    if all_sheet is None:
        issues.append({"code": "all_reviews_sheet_missing", "sheet": all_spec.get("sheet")})
    if classified_sheet is None:
        issues.append({"code": "classified_reviews_sheet_missing", "sheet": classified_spec.get("sheet")})

    all_columns: dict[str, int] = {}
    classified_columns: dict[str, int] = {}
    subcategory_indices: dict[str, int] = {}
    if all_sheet is not None:
        all_columns, missing = header_map(all_sheet, all_spec["columns"])
        if missing:
            issues.append({"code": "all_reviews_columns_missing", "headers": missing})
    if classified_sheet is not None:
        requested = dict(classified_spec["columns"])
        requested.update({f"subcategory::{category}": header for category, header in config["classification"]["subcategory_columns"].items()})
        classified_columns, missing = header_map(classified_sheet, requested)
        if missing:
            issues.append({"code": "classified_reviews_columns_missing", "headers": missing})
        subcategory_indices = {
            category: classified_columns.get(f"subcategory::{category}", -1)
            for category in config["classification"]["categories"]
        }

    start = date.fromisoformat(config["period"]["start"]) if not any(i["code"] == "config_period_invalid" for i in issues) else None
    end = date.fromisoformat(config["period"]["end"]) if start else None
    all_rows = []
    detail_rows = []

    if all_sheet is not None and len(all_columns) == len(all_spec["columns"]):
        seen = Counter()
        invalid_rows = []
        outside_rows = []
        dates = []
        for row_index, row in enumerate(all_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            seen[exact_row_key(row)] += 1
            parsed = parse_date(row[all_columns["date"]], wb.epoch)
            try:
                rating = int(row[all_columns["rating"]])
                rating_ok = float(row[all_columns["rating"]]) == rating and 1 <= rating <= 5
            except Exception:
                rating, rating_ok = None, False
            if parsed is None or not rating_ok:
                invalid_rows.append(row_index)
                continue
            dates.append(parsed)
            if start and end and not start <= parsed <= end:
                outside_rows.append(row_index)
            all_rows.append({
                "row": row_index,
                "date": parsed.isoformat(),
                "version": str(row[all_columns["version"]] or "").strip(),
                "rating": rating,
                "review": str(row[all_columns["review"]] or "").strip(),
            })
        duplicate_count = sum(count - 1 for count in seen.values() if count > 1)
        if duplicate_count:
            issues.append({"code": "exact_duplicate_rows", "count": duplicate_count})
        if invalid_rows:
            issues.append({"code": "invalid_date_or_rating", "rows": invalid_rows[:50], "count": len(invalid_rows)})
        if outside_rows:
            issues.append({"code": "rows_outside_configured_period", "rows": outside_rows[:50], "count": len(outside_rows)})
        if not dates:
            issues.append({"code": "no_valid_all_review_rows"})
        elif config["period"].get("require_every_date") and start and end:
            present = set(dates)
            expected = {start + timedelta(days=i) for i in range((end - start).days + 1)}
            missing = sorted(value.isoformat() for value in expected - present)
            if missing:
                issues.append({"code": "period_dates_missing", "dates": missing})

    if classified_sheet is not None and all(index >= 0 for index in subcategory_indices.values()):
        categories = set(config["classification"]["categories"])
        invalid_rows = []
        for row_index, row in enumerate(classified_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            parsed = parse_date(row[classified_columns["date"]], wb.epoch)
            try:
                rating = int(row[classified_columns["rating"]])
            except Exception:
                rating = None
            category = str(row[classified_columns["category"]] or "").strip()
            original = str(row[classified_columns["original"]] or "").strip()
            translation = str(row[classified_columns["translation"]] or "").strip()
            subcategory = str(row[subcategory_indices.get(category, -1)] or "").strip() if category in subcategory_indices else ""
            if parsed is None or category not in categories or not subcategory or not original or rating not in config["metrics"]["low_ratings"]:
                invalid_rows.append(row_index)
            detail_rows.append({
                "row": row_index,
                "date": parsed.isoformat() if parsed else None,
                "rating": rating,
                "category": category,
                "subcategory": subcategory,
                "original": original,
                "translation": translation,
            })
        if invalid_rows:
            issues.append({"code": "classified_row_invalid", "rows": invalid_rows[:50], "count": len(invalid_rows)})

    if all_rows and detail_rows:
        low_all = Counter(
            (row["date"], row["rating"], row["review"])
            for row in all_rows
            if row["rating"] in config["metrics"]["low_ratings"] and row["review"]
        )
        low_detail = Counter((row["date"], row["rating"], row["original"]) for row in detail_rows)
        if low_all != low_detail:
            issues.append({
                "code": "classified_coverage_mismatch",
                "all_review_count": sum(low_all.values()),
                "classified_count": sum(low_detail.values()),
                "missing_from_classified": sum((low_all - low_detail).values()),
                "extra_in_classified": sum((low_detail - low_all).values()),
            })

    result = {
        "status": "NEEDS_USER" if issues else "PASS",
        "needs_user": bool(issues),
        "workbook": str(workbook_path),
        "previous_report": str(previous_path),
        "config": str(config_path),
        "all_reviews": {"sheet": all_spec.get("sheet"), "rows": all_rows},
        "classified_reviews": {"sheet": classified_spec.get("sheet"), "rows": detail_rows},
        "period": {"start": config["period"]["start"], "end": config["period"]["end"]},
        "issues": issues,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "status": result["status"],
        "period": result["period"],
        "all_rows": len(all_rows),
        "classified_rows": len(detail_rows),
        "issues": issues,
    }, ensure_ascii=False))
    raise SystemExit(2 if issues else 0)


if __name__ == "__main__":
    main()

